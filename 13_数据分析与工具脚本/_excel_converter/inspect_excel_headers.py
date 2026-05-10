#!/usr/bin/env python3
# =============================================================================
# inspect_excel_headers.py - xlsx 字段结构探测工具 v0.1
# =============================================================================
# 用途：对 convert_xls_to_xlsx.sh 转换后的 xlsx 做 header 结构探测
# 仅读 sheet 名 / max_row / max_column / 前 15 行非空摘要 / header 候选
# 不输出真实大段明细数据 / 不写入结果文件 / 不修改原文件
#
# 使用：
#   python3 inspect_excel_headers.py <xlsx 目录>
#
# 例：
#   python3 inspect_excel_headers.py /tmp/_xls_convert_20260510
# =============================================================================

import os
import sys
import argparse
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    print("❌ 需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# 探测前若干行 / 找 header
TOP_ROWS_DEFAULT = 15

# 单行最大显示列数（防止一行字段太多刷屏）
MAX_COLS_DISPLAY = 30


def inspect_one_xlsx(fpath, top_rows=TOP_ROWS_DEFAULT):
    """探测一个 xlsx 文件 / 返回结构 dict / 无副作用 / 不修改文件"""
    result = {
        "file": os.path.basename(fpath),
        "size_mb": round(os.path.getsize(fpath) / 1024 / 1024, 2),
        "sheets": [],
        "error": None,
    }
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            sheet_info = {
                "name": sn,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "top_rows": [],
                "header_candidate_row": None,
                "header_candidate_fields": [],
            }
            top = []
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=top_rows, values_only=True), 1):
                cleaned = [str(c).strip() if c is not None else "" for c in row]
                top.append(cleaned)
            sheet_info["top_rows"] = top

            # header 候选 = 前 top_rows 内最大非空列数那行
            if top:
                best_idx = 0
                best_count = -1
                for i, r in enumerate(top):
                    cnt = sum(1 for c in r if c)
                    if cnt > best_count:
                        best_count = cnt
                        best_idx = i
                sheet_info["header_candidate_row"] = best_idx + 1  # 1-based
                sheet_info["header_candidate_fields"] = top[best_idx]

            result["sheets"].append(sheet_info)
        wb.close()
    except Exception as e:
        result["error"] = str(e)
    return result


def print_one_result(r):
    """打印一个文件的探测结果 / 不输出真实数据明细"""
    print(f"\n📄 {r['file']}  ({r['size_mb']} MB)")
    if r["error"]:
        print(f"   ❌ 读取失败: {r['error']}")
        return
    for sh in r["sheets"]:
        print(f"   📋 [{sh['name']}]  {sh['max_row']} 行 × {sh['max_col']} 列")
        if sh["header_candidate_row"]:
            hr = sh["header_candidate_row"]
            fields = sh["header_candidate_fields"]
            non_empty = [f for f in fields if f]
            print(f"      🎯 header 候选 (Row {hr}, {len(non_empty)} 非空字段):")
            # 只显示非空字段（带原列号）
            for i, f in enumerate(fields):
                if f and i < MAX_COLS_DISPLAY:
                    print(f"         col {i+1}: {f}")
            if len([f for f in fields if f]) > MAX_COLS_DISPLAY:
                rest = [(i + 1, f) for i, f in enumerate(fields) if f][MAX_COLS_DISPLAY:]
                print(f"         ... 还有 {len(rest)} 列（已截断）")
        else:
            print(f"      ⚠️  无 header 候选（前 {TOP_ROWS_DEFAULT} 行全空）")
        # 非 header 行的非空摘要
        if sh["top_rows"]:
            print(f"      📑 前 {len(sh['top_rows'])} 行结构概览:")
            for ri, row in enumerate(sh["top_rows"], 1):
                non_empty_count = sum(1 for c in row if c)
                if non_empty_count == 0:
                    summary = "(空行)"
                elif non_empty_count <= 5:
                    # 全部展示（只有几个字段）
                    summary = " | ".join(f"col{i+1}={c}" for i, c in enumerate(row) if c)
                else:
                    summary = f"({non_empty_count} 非空字段)"
                print(f"         R{ri:2d}: {summary}")


def main():
    parser = argparse.ArgumentParser(
        description="xlsx 字段结构探测 / 仅读 header / 不输出明细数据 / 不修改文件"
    )
    parser.add_argument("input_dir", help="xlsx 文件目录（通常是 convert_xls_to_xlsx.sh 输出目录）")
    parser.add_argument("--top-rows", type=int, default=TOP_ROWS_DEFAULT,
                        help=f"探测前 N 行寻找 header 候选（默认 {TOP_ROWS_DEFAULT}）")
    args = parser.parse_args()

    input_dir = args.input_dir
    if not os.path.isdir(input_dir):
        print(f"❌ 输入目录不存在: {input_dir}", file=sys.stderr)
        sys.exit(1)

    xlsx_files = sorted([
        f for f in os.listdir(input_dir)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ])

    if not xlsx_files:
        print(f"⚠️  未在 {input_dir} 找到 .xlsx 文件")
        sys.exit(0)

    print("=" * 100)
    print(f"📋 xlsx 字段结构探测 v0.1")
    print(f"   目录: {input_dir}")
    print(f"   文件数: {len(xlsx_files)}")
    print(f"   探测前 {args.top_rows} 行 / 仅 header 结构 / 不输出明细数据")
    print("=" * 100)

    success = 0
    failed = 0
    for fname in xlsx_files:
        fpath = os.path.join(input_dir, fname)
        r = inspect_one_xlsx(fpath, top_rows=args.top_rows)
        print_one_result(r)
        if r["error"]:
            failed += 1
        else:
            success += 1

    print("\n" + "=" * 100)
    print(f"✅ 完成: {success} 成功 / {failed} 失败 / 共 {len(xlsx_files)} 文件")
    print("✅ 仅 header 结构 / 0 数据明细输出")
    print("=" * 100)


if __name__ == "__main__":
    main()
