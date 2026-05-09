"""
只读 Excel 加载层。
绝不调用 .save() / .write() 至原始路径。
"""
import hashlib
from pathlib import Path
from typing import Iterator


def compute_sha256(path: Path) -> str:
    """计算文件 SHA256，只读 buffer。"""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_xls_readonly(path: Path) -> dict:
    """旧 .xls 格式 — xlrd（默认无写入能力）。"""
    import xlrd
    book = xlrd.open_workbook(str(path), on_demand=True)
    sheets = {}
    for sheet_name in book.sheet_names():
        sh = book.sheet_by_name(sheet_name)
        rows = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                val = sh.cell_value(r, c)
                # xlrd 数字默认 float；空 cell 为 ""
                row.append(val)
            rows.append(row)
        sheets[sheet_name] = {
            "rows": rows,
            "nrows": sh.nrows,
            "ncols": sh.ncols,
        }
        book.unload_sheet(sheet_name)
    return sheets


def load_xlsx_readonly(path: Path) -> dict:
    """新 .xlsx 格式 — openpyxl read_only=True 强制只读。"""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[sheet_name] = {
            "rows": rows,
            "nrows": len(rows),
            "ncols": max((len(r) for r in rows), default=0),
        }
    wb.close()  # 注意：openpyxl read_only 模式下 close() 不写入
    return sheets


def load_excel_readonly(path: Path) -> dict:
    """统一只读加载 .xls / .xlsx。"""
    if not path.exists():
        raise FileNotFoundError(f"输入文件不存在: {path}")
    ext = path.suffix.lower()
    if ext == ".xls":
        return load_xls_readonly(path)
    elif ext == ".xlsx":
        return load_xlsx_readonly(path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def verify_no_modification(input_files: dict, sha_before: dict) -> dict:
    """重新计算 SHA256 与执行前对比，确认零修改。"""
    result = {"all_match": True, "mismatches": []}
    for fid, path in input_files.items():
        sha_after = compute_sha256(path)
        if sha_before.get(fid) != sha_after:
            result["all_match"] = False
            result["mismatches"].append({
                "fid": fid,
                "before": sha_before.get(fid),
                "after": sha_after,
            })
    return result
